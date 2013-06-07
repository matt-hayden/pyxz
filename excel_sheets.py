#!env python
from collections import namedtuple
from logging import debug, info, warning, error, critical

from openpyxl import load_workbook

from TraceWizard4.MDB_File import sql_field_sanitize

def gen_table_from_XLSX(filename = '',		# (or workbook keyword arg)
						sheetname = '',		# also accessible as sheet.title
						extent = '',		# (rectangular?) range of table
						header = True,		# True for import header, False for don't, any other value sets the header
						**kwargs):
	assert filename or (workbook in kwargs)
	nz = kwargs.pop('NZ', lambda x: x or 0)
	workbook = kwargs.pop('workbook', None) or load_workbook(filename)
	sheet = workbook.get_sheet_by_name(sheetname) if sheetname else workbook.get_active_sheet()
	debug("Using sheet {} out of {}".format(sheet.title, workbook.get_sheet_names()))
	extent = extent or sheet.calculate_dimension()
	info("Using extent {}".format(extent))
	range = sheet.range(extent)
	if header:
		if header is True:
			header = [_.value for _ in range[0]]
			range = range[1:]
		factory = kwargs.pop('factory', None) or namedtuple('Row', [sql_field_sanitize(_) for _ in header])
		for row in range:
			yield factory(*[nz(_.value) for _ in row])
	elif header in (False, ''):
		for row in range:
			yield [nz(_.value) for _ in row]
#
if __name__ == '__main__':
	import csv
	import sys
	
	from console_size import to_columns
	from myglob import glob
	#
	args = sys.argv[1:]
	nargs = len(args)
	if nargs > 2:
		filename, sheetnames = args[0], args[1:]
		wb = load_workbook(filename)
		for s in sheetnames:
			print s
			for row in gen_table_from_XLSX(workbook=wb, sheetname=s):
				print '\t'.join(row)
	elif nargs == 1:
		(arg,) = args
		wb = load_workbook(arg)
		print "[{}]:".format(arg)
		print to_columns(wb.get_sheet_names())
	else:
		args = glob('*.XLSX')
		if args:
			width = max(len(_) for _ in args)
			for arg in args:
				wb = load_workbook(arg)
				print arg.ljust(width)+":"+str(wb.get_sheet_names())
			print
		else:
			print >>sys.stderr, "No files specified"
			sys.exit(-1)