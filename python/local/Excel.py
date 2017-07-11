#!env python
"""Working with XLS and XLSX files using third-party libraries.


"""
from collections import namedtuple

import openpyxl
import xlrd

from sanitize import sql_field_sanitize

def gen_table_from_XLSX(filename = '',		# (or workbook keyword arg)
						sheetname = '',		# also accessible as sheet.title, defaults to last active sheet
						extent = '',		# (rectangular?) range of table
						header = True,		# True for import header, False for don't, any other value sets the header
						**kwargs):
	assert filename or (workbook in kwargs)
	nz = kwargs.pop('NZ', lambda x: x or 0)
	workbook = kwargs.pop('workbook', None) or openpyxl.load_workbook(filename, use_iterators=True)
	sheet = workbook.get_sheet_by_name(sheetname) if sheetname else workbook.get_active_sheet()
	debug("Using sheet {} out of {}".format(sheet.title, workbook.get_sheet_names()))
#	extent = extent or sheet.calculate_dimension()
#	info("Using extent {}".format(extent))
#	crange = sheet.range(extent)
	crange = sheet.iter_rows()
	if header:
		if header is True:
#			header = [c.value for c in crange[0]]
			header = [c.internal_value for c in crange.next()]
#			crange = crange[1:]
		factory = kwargs.pop('factory', None) or namedtuple('Row', [sql_field_sanitize(f) for f in header])
		for row in crange:
#			yield factory(*[nz(c.value) for c in row])
			yield factory(*[nz(c.internal_value) for c in row])
	elif header in (False, ''):
		for row in crange:
#			yield [nz(c.value) for c in row]
			yield [nz(c.internal_value) for c in row]
#
def gen_table_from_XLS(filename = '',		# (or workbook keyword arg)
					   sheetname = '',		# also accessible as sheet.name, defaults to leftmost sheet
					   header = True,		# True for import header, False for don't, any other value sets the header
					   **kwargs):
	assert filename or (workbook in kwargs)
	nz = kwargs.pop('NZ', lambda x: x or 0)
	with kwargs.pop('workbook', None) or xlrd.open_workbook(filename) as workbook:
		sheet = workbook.sheet_by_name(sheetname) if sheetname else workbook.sheet_by_index(0)
		debug("Using sheet {} out of {}".format(sheet.name, workbook.sheet_names()))
		#
		start, end = 0, sheet.nrows
		if header:
			if header is True:
				header = [c.value for c in sheet.row_slice(0)]
				start += 1
			factory = kwargs.pop('factory', None) or namedtuple('Row', [sql_field_sanitize(f) for f in header])
			for r in xrange(start, end):
				yield factory(*[nz(c.value) for c in sheet.row_slice(r)])
		elif header in (False, ''):
			for r in xrange(start, end):
				yield [nz(c.value) for c in sheet.row_slice(r)]
#
def get_sheet_names_XLSX(*args, **kwargs):
	wb = openpyxl.load_workbook(*args, **kwargs)
	return wb.get_sheet_names()
#
def get_sheet_names_XLS(*args, **kwargs):
	with xlrd.open_workbook(*args, **kwargs) as workbook:
		return workbook.sheet_names()
#
if __name__ == '__main__':
	import csv
	import sys
	
	from console_size import to_columns
	from myglob import glob
	#
	args = sys.argv[1:]
	nargs = len(args)
	if nargs > 2:
		filename, sheetnames = args[0], args[1:]
		wb = openpyxl.load_workbook(filename)
		for s in sheetnames:
			print s
			for row in gen_table_from_XLSX(workbook=wb, sheetname=s):
				print '\t'.join(row)
	elif nargs == 1:
		(arg,) = args
		wb = openpyxl.load_workbook(arg)
		print "[{}]:".format(arg)
		print to_columns(wb.get_sheet_names())
	else:
		args = glob('*.XLSX')
		if args:
			width = max(len(_) for _ in args)
			for arg in args:
				wb = openpyxl.load_workbook(arg)
				print arg.ljust(width)+":"+str(wb.get_sheet_names())
			print
		else:
			print >>sys.stderr, "No files specified"
			sys.exit(-1)